import pandas as pd
import numpy as np
import ANOVA
import TukeyHSD
from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import Font
from openpyxl.styles import Font
import TreatmentName

excel_file_path = "resources/TomatoData.xlsx"
workbook = load_workbook(excel_file_path)

sheet = workbook["Yield components"]

density_from_spacing = [0, 66700, 57100, 50000, 40000]
density_from_spacing = [x / 1000000 for x in density_from_spacing]


def readExcelSheet(sheet):
    data = []
    color = []
    for row in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
    ):
        data_row = []
        color_row = []
        for cell in row:
            data_row.append(cell.value)
            text_color = cell.font.color
            if isinstance(text_color, Color):
                if text_color.rgb is None:
                    text_color_str = "Default"
                else:
                    text_color_str = text_color.rgb
            else:
                text_color_str = "Default"
            color_row.append(text_color_str)

        data.append(data_row)
        color.append(color_row)
    df_data = pd.DataFrame(data[1:], columns=data[0])
    # print(color)
    df_color = pd.DataFrame(color[1:], columns=data[0])
    return df_data, df_color


df_data, df_color = readExcelSheet(sheet)


def experimentCellsToNpArray(df_data, df_color, numExperiments=5, maxTruss=4):
    res = pd.DataFrame(
        columns=[
            "Replication",
            "Treatment",
            "Spacing",
            "Truss",
            "Number of\nfruit",
            "Average fruit\nweight",
            "Individual\nFruit Yield",
            "Fruit Yield",
            "Number of\nmarketable fruit",
            "Average marketable\nfruit weight",
            "Individual Marketable\nFruit Yield",
            "Marketable\nFruit Yield",
            "Marketable Fruit\nNumber 2Ws",
            "Marketable Fruit\nWeight 2Ws",
            "Individual Marketable\nFruit Yield 2Ws",
            "Marketable Fruit\nYield 2Ws",
            "Yield Truss 1",
            "Yield Truss 2",
            "Yield Truss 3",
            "Yield Truss 4",
            "Marketable Yield\nTruss 1",
            "Marketable Yield\nTruss 2",
            "Marketable Yield\nTruss 3",
            "Marketable Yield\nTruss 4",
        ]
    )

    # Loop through all rows
    for index, row in df_data.iterrows():
        treatment = row.iat[0]
        if TreatmentName.check(treatment):
            replication, Spacing, Truss = TreatmentName.standard3(treatment)
            treatment = TreatmentName.standard(treatment)[1]
            nPlants = density_from_spacing[Spacing]

            # count next rows
            start_index = index + 1
            end_index = start_index
            while end_index == start_index or pd.isna(df_data.iat[end_index, 0]):
                end_index += 1
            # [start_index, end_index) is the range of data

            numTruss = Truss + 1
            (trussYield, trussMarketable) = (
                [[] for _ in range(maxTruss)] for _ in range(2)
            )

            (
                cnt_plant,
                sum_plant,
                cnt_marketable_plant,
                sum_marketable_plant,
                cnt_marketable_plant_2w,
                sum_marketable_plant_2w,
            ) = ([] for _ in range(6))

            def check2w(color):
                return color != "FF00B0F0"

            cur_col = 0
            for expId in range(numExperiments):
                (total, marketable, marketable_2w) = ([] for _ in range(3))
                for trussId in range(0, numTruss):
                    trussYield_plant = 0
                    trussMarketable_plant = 0

                    cur_col += 1
                    for k in range(start_index, end_index):
                        if pd.notna(df_data.iat[k, cur_col]):
                            trussYield_plant += df_data.iat[k, cur_col]

                            total.append(df_data.iat[k, cur_col])

                            if df_data.iat[k, cur_col] > 60:
                                trussMarketable_plant += df_data.iat[k, cur_col]

                                marketable.append(df_data.iat[k, cur_col])

                                if check2w(df_color.iat[k, cur_col]):
                                    marketable_2w.append(df_data.iat[k, cur_col])

                    trussYield[trussId].append(trussYield_plant * nPlants)
                    trussMarketable[trussId].append(trussMarketable_plant * nPlants)
                if len(total) > 0:
                    cnt_plant.append(len(total))
                    sum_plant.append(np.sum(total))
                    cnt_marketable_plant.append(len(marketable))
                    sum_marketable_plant.append(np.sum(marketable))
                    cnt_marketable_plant_2w.append(len(marketable_2w))
                    sum_marketable_plant_2w.append(np.sum(marketable_2w))

            nPlants = density_from_spacing[Spacing]

            Number_of_fruit = ANOVA.calc_mean_sstot_size(cnt_plant)
            Average_fruit_weight = ANOVA.calc_mean_sstot_size(
                [x / y if y != 0 else 0 for x, y in zip(sum_plant, cnt_plant)]
            )
            Individual_Fruit_Yield = ANOVA.calc_mean_sstot_size(sum_plant)
            Fruit_Yield = [
                Individual_Fruit_Yield[0] * nPlants,
                Individual_Fruit_Yield[1] * (nPlants**2),
                Individual_Fruit_Yield[2],
            ]

            Number_of_marketable_fruit = ANOVA.calc_mean_sstot_size(
                cnt_marketable_plant
            )
            Average_marketable_fruit_weight = ANOVA.calc_mean_sstot_size(
                [
                    x / y if y != 0 else 0
                    for x, y in zip(sum_marketable_plant, cnt_marketable_plant)
                ]
            )
            Individual_Marketable_Fruit_Yield = ANOVA.calc_mean_sstot_size(
                sum_marketable_plant
            )
            Marketable_Fruit_Yield = [
                Individual_Marketable_Fruit_Yield[0] * nPlants,
                Individual_Marketable_Fruit_Yield[1] * (nPlants**2),
                Individual_Marketable_Fruit_Yield[2],
            ]

            Marketable_Fruit_Number_2Ws = ANOVA.calc_mean_sstot_size(
                cnt_marketable_plant_2w
            )
            Marketable_Fruit_Weight_2Ws = ANOVA.calc_mean_sstot_size(
                [
                    x / y if y != 0 else 0
                    for x, y in zip(sum_marketable_plant_2w, cnt_marketable_plant_2w)
                ]
            )
            Individual_Marketable_Fruit_Yield_2Ws = ANOVA.calc_mean_sstot_size(
                sum_marketable_plant_2w
            )
            Marketable_Fruit_Yield_2Ws = [
                Individual_Marketable_Fruit_Yield_2Ws[0] * nPlants,
                Individual_Marketable_Fruit_Yield_2Ws[1] * (nPlants**2),
                Individual_Marketable_Fruit_Yield_2Ws[2],
            ]

            trussYield = [
                ANOVA.calc_mean_sstot_size(x) if len(x) > 0 else [np.nan] * 3
                for x in trussYield
            ]
            trussMarketable = [
                ANOVA.calc_mean_sstot_size(x) if len(x) > 0 else [np.nan] * 3
                for x in trussMarketable
            ]

            res.loc[len(res)] = (
                replication,
                treatment,
                Spacing,
                Truss,
                Number_of_fruit,
                Average_fruit_weight,
                Individual_Fruit_Yield,
                Fruit_Yield,
                Number_of_marketable_fruit,
                Average_marketable_fruit_weight,
                Individual_Marketable_Fruit_Yield,
                Marketable_Fruit_Yield,
                Marketable_Fruit_Number_2Ws,
                Marketable_Fruit_Weight_2Ws,
                Individual_Marketable_Fruit_Yield_2Ws,
                Marketable_Fruit_Yield_2Ws,
                trussYield[0],
                trussYield[1],
                trussYield[2],
                trussYield[3],
                trussMarketable[0],
                trussMarketable[1],
                trussMarketable[2],
                trussMarketable[3],
            )
    return res


df_data, df_color = readExcelSheet(sheet)
avg = experimentCellsToNpArray(df_data, df_color)


def getAvgAllReplication(avg):
    tmp = avg.drop(columns=["Replication", "Treatment"], axis=1)
    res = pd.DataFrame(columns=tmp.columns)
    for (spacing, truss), group_data in tmp.groupby(["Spacing", "Truss"]):
        row = pd.Series()
        row["Spacing"] = spacing
        row["Truss"] = truss
        for col in group_data.columns:
            if col not in ["Spacing", "Truss"]:
                groups_mean = [x[0] for x in group_data[col]]
                groups_ssw = sum([x[1] for x in group_data[col]])
                groups_size = [x[2] for x in group_data[col]]

                mean, ssw, size = ANOVA.summary_data(
                    groups_mean, groups_ssw, groups_size
                )
                # print(col, mean, ssw, size)
                row[col] = [mean, ssw, size]
        res = res._append(row, ignore_index=True)
    # print(res)
    return res


tmp = getAvgAllReplication(avg)

cols_to_show = [
    "Number of\nfruit",
    "Average fruit\nweight",
    "Individual\nFruit Yield",
    "Fruit Yield",
    "Number of\nmarketable fruit",
    "Average marketable\nfruit weight",
    "Individual Marketable\nFruit Yield",
    "Marketable\nFruit Yield",
]


def groupSpacing(df):
    df = df[["Spacing"] + cols_to_show]
    res = pd.DataFrame(columns=df.columns)
    return ANOVA.ANOVA_test_summary_table(df, res, "Spacing")


def groupTruss(df):
    df = df[["Truss"] + cols_to_show]
    res = pd.DataFrame(columns=df.columns)
    return ANOVA.ANOVA_test_summary_table(df, res, "Truss")


def groupTreatment(df):
    df = df[["Treatment"] + cols_to_show]
    res = pd.DataFrame(columns=df.columns)
    return ANOVA.ANOVA_test_summary_table(df, res, "Treatment")


def groupTreatmentnoANOVA(df):
    res = pd.DataFrame(columns=df.columns)
    res.drop(columns=["Replication", "Spacing", "Truss"], inplace=True)
    return ANOVA.ANOVA_test_summary_table(df, res, "Treatment", ANOVAtest=False)[
        [
            "Treatment",
            "Fruit Yield",
            "Marketable\nFruit Yield",
            "Marketable Fruit\nYield 2Ws",
        ]
    ]


def groupTreatmentTrussnoANOVA(df):
    res = pd.DataFrame(columns=df.columns)
    res.drop(columns=["Replication", "Spacing", "Truss"], inplace=True)
    return ANOVA.ANOVA_test_summary_table(
        df, res, "Treatment", MSEprecision=2, ANOVAtest=False
    )[
        [
            "Treatment",
            "Yield Truss 1",
            "Yield Truss 2",
            "Yield Truss 3",
            "Yield Truss 4",
            "Marketable Yield\nTruss 1",
            "Marketable Yield\nTruss 2",
            "Marketable Yield\nTruss 3",
            "Marketable Yield\nTruss 4",
        ]
    ]


def groupMarketableYield2W(df):
    unique_spacing = df["Spacing"].unique()
    unique_truss = df["Truss"].unique()
    df["Treatment"] = "S" + df["Spacing"].astype(str) + "T" + df["Truss"].astype(str)

    df = df[["Treatment", "Marketable Fruit\nYield 2Ws"]]
    tmp = pd.DataFrame(columns=df.columns)
    tmp = ANOVA.ANOVA_test_summary_table(df, tmp, "Treatment")

    res = pd.DataFrame(index=unique_spacing, columns=unique_truss)
    res2 = pd.DataFrame(index=unique_spacing, columns=unique_truss)

    for spacing in unique_spacing:
        for truss in unique_truss:
            res.at[spacing, truss] = tmp[(tmp["Treatment"] == f"S{spacing}T{truss}")][
                "Marketable Fruit\nYield 2Ws"
            ].values
            res2.at[spacing, truss] = df[(df["Treatment"] == f"S{spacing}T{truss}")][
                "Marketable Fruit\nYield 2Ws"
            ].values[0]

    tmpS = [
        ANOVA.summary_data(
            [res2.at[spacing, truss][0] for truss in unique_truss],
            sum(res2.at[spacing, truss][1] for truss in unique_truss),
            [res2.at[spacing, truss][2] for truss in unique_truss],
        )
        for spacing in unique_spacing
    ]
    meanS = [val[0] for val in tmpS]
    sswS = sum([val[1] for val in tmpS])
    sizeS = [val[2] for val in tmpS]

    lsdS, charGroupS = ANOVA.perform_anova_summary(meanS, sswS, sizeS, alpha=0.05)

    for i in range(len(meanS)):
        meanS[i] = str(round(meanS[i], 2)) + TukeyHSD.get_superscript(charGroupS[i])

    tmpT = [
        ANOVA.summary_data(
            [res2.at[spacing, truss][0] for spacing in unique_spacing],
            sum(res2.at[spacing, truss][1] for spacing in unique_spacing),
            [res2.at[spacing, truss][2] for spacing in unique_spacing],
        )
        for truss in unique_truss
    ]

    meanT = [val[0] for val in tmpT]
    sswT = sum([val[1] for val in tmpT])
    sizeT = [val[2] for val in tmpT]

    lsdT, charGroupT = ANOVA.perform_anova_summary(meanT, sswT, sizeT, alpha=0.05)

    for i in range(len(meanT)):
        meanT[i] = str(round(meanT[i], 2)) + TukeyHSD.get_superscript(charGroupT[i])

    res.index = ["40 × 30", "40 × 35", "40 × 40", "40 × 50"]
    res.columns = ["2 trusses", "3 trusses", "4 trusses"]

    res["Mean"] = meanS
    res.loc["Mean"] = meanT + [""]
    lsd = "LSD" + TukeyHSD.get_subscript("(0.05)")
    res[lsd] = [""] * len(unique_spacing) + [lsdS]
    res.loc[lsd] = [""] * len(unique_truss) + [lsdT, ""]
    # print(res)
    return res


Table2 = groupSpacing(tmp)
Table3 = groupTruss(tmp)
Table4 = groupTreatment(avg)
Table5 = groupTreatmentTrussnoANOVA(avg)
Table6 = groupMarketableYield2W(tmp)
Fig2Table = groupTreatmentnoANOVA(avg)
