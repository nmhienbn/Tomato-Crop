import pandas as pd
import numpy as np
import ANOVA
import TukeyHSD
from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import Font
from openpyxl.styles import Font
import TreatmentName

excel_file_path = "resources/TomatoData.xlsx"
tableDir = "outputs/Table2_Yield_Component.png"
workbook = load_workbook(excel_file_path)

sheet = workbook["Yield components"]

density_from_spacing = [0, 66700, 57100, 50000, 40000]
density_from_spacing = [x / 1000000 for x in density_from_spacing]


def readExcelSheet(sheet):
    data = []
    color = []
    for row in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
    ):
        data_row = []
        color_row = []
        for cell in row:
            data_row.append(cell.value)
            text_color = cell.font.color
            if isinstance(text_color, Color):
                if text_color.rgb is None:
                    text_color_str = "Default"
                else:
                    text_color_str = text_color.rgb
            else:
                text_color_str = "Default"
            color_row.append(text_color_str)

        data.append(data_row)
        color.append(color_row)
    df_data = pd.DataFrame(data[1:], columns=data[0])
    # print(color)
    df_color = pd.DataFrame(color[1:], columns=data[0])
    return df_data, df_color


df_data, df_color = readExcelSheet(sheet)


def experimentCellsToNpArray(df_data, df_color, numExperiments=5):
    res = pd.DataFrame(
        columns=[
            "Replication",
            "Treatment",
            "Spacing",
            "Truss",
            # "Truss 1",
            # "Truss 2",
            # "Truss 3",
            # "Truss 4",
            # "Truss 5",
            "Number of fruit",
            "Average fruit weight",
            "Individual Fruit Yield",
            "Fruit Yield",
            "Number of marketable fruit",
            "Average marketable fruit weight",
            "Individual Marketable Fruit Yield",
            "Marketable Fruit Yield",
            "Marketable Fruit Number 2Ws",
            "Marketable Fruit Weight 2Ws",
            "Individual Marketable Fruit Yield 2Ws",
            "Marketable Fruit Yield 2Ws",
            # "Marketable Yield 2Ws",
        ]
    )

    # Loop through all rows
    for index, row in df_data.iterrows():
        treatment = row.iat[0]
        if TreatmentName.check(treatment):
            replication, Spacing, Truss = TreatmentName.standard3(treatment)
            treatment = TreatmentName.standard(treatment)[1]

            # count next rows
            start_index = index + 1
            end_index = start_index
            while end_index == start_index or pd.isna(df_data.iat[end_index, 0]):
                end_index += 1
            # [start_index, end_index) is the range of data

            numTruss = Truss + 1
            cnt = [0] * numTruss
            cnt_marketable = [0] * numTruss
            sum = [0] * numTruss
            sum_marketable = [0] * numTruss

            cntExperiments = 0
            for i in range(numExperiments):
                if pd.notna(df_data.iat[start_index, i * numTruss + 1]):
                    cntExperiments += 1

            cnt_plant = [0] * cntExperiments
            cnt_marketable_plant = [0] * cntExperiments

            sum_plant = [0] * cntExperiments
            sum_marketable_plant = [0] * cntExperiments

            cnt_marketable_2w = [0] * cntExperiments
            sum_marketable_2w = [0] * cntExperiments

            def check2w(color):
                return color != "FF00B0F0"

            cur_col = 0
            for i in range(cntExperiments):
                for j in range(0, numTruss):
                    cur_col += 1
                    for k in range(start_index, end_index):
                        if pd.notna(df_data.iat[k, cur_col]):
                            sum[j] += df_data.iat[k, cur_col]
                            cnt[j] += 1

                            sum_plant[i] += df_data.iat[k, cur_col]
                            cnt_plant[i] += 1

                            if df_data.iat[k, cur_col] > 60:
                                sum_marketable[j] += df_data.iat[k, cur_col]
                                cnt_marketable[j] += 1

                                sum_marketable_plant[i] += df_data.iat[k, cur_col]
                                cnt_marketable_plant[i] += 1

                                if check2w(df_color.iat[k, cur_col]):
                                    sum_marketable_2w[i] += df_data.iat[k, cur_col]
                                    cnt_marketable_2w[i] += 1

            nPlants = density_from_spacing[Spacing]

            Number_of_fruit = ANOVA.calc_mean_sstot_size(cnt_plant)
            Average_fruit_weight = ANOVA.calc_mean_sstot_size(
                [x / y for x, y in zip(sum_plant, cnt_plant) if y != 0]
            )
            Individual_Fruit_Yield = ANOVA.calc_mean_sstot_size(sum_plant)
            Fruit_Yield = [
                Individual_Fruit_Yield[0] * nPlants,
                Individual_Fruit_Yield[1] * (nPlants**2),
                Individual_Fruit_Yield[2],
            ]

            Number_of_marketable_fruit = ANOVA.calc_mean_sstot_size(
                cnt_marketable_plant
            )
            Average_marketable_fruit_weight = ANOVA.calc_mean_sstot_size(
                [
                    x / y
                    for x, y in zip(sum_marketable_plant, cnt_marketable_plant)
                    if y != 0
                ]
            )
            Individual_Marketable_Fruit_Yield = ANOVA.calc_mean_sstot_size(
                sum_marketable_plant
            )
            Marketable_Fruit_Yield = [
                Individual_Marketable_Fruit_Yield[0] * nPlants,
                Individual_Marketable_Fruit_Yield[1] * (nPlants**2),
                Individual_Marketable_Fruit_Yield[2],
            ]

            Marketable_Fruit_Number_2Ws = ANOVA.calc_mean_sstot_size(cnt_marketable_2w)
            Marketable_Fruit_Weight_2Ws = ANOVA.calc_mean_sstot_size(
                [x / y for x, y in zip(sum_marketable_2w, cnt_marketable_2w) if y != 0]
            )
            Individual_Marketable_Fruit_Yield_2Ws = ANOVA.calc_mean_sstot_size(
                sum_marketable_2w
            )
            Marketable_Fruit_Yield_2Ws = [
                Individual_Marketable_Fruit_Yield_2Ws[0] * nPlants,
                Individual_Marketable_Fruit_Yield_2Ws[1] * (nPlants**2),
                Individual_Marketable_Fruit_Yield_2Ws[2],
            ]

            def roundl(x, n):
                return [round(y, n) for y in x]

            res.loc[len(res)] = (
                replication,
                treatment,
                Spacing,
                Truss,
                roundl(Number_of_fruit, 2),
                roundl(Average_fruit_weight, 2),
                roundl(Individual_Fruit_Yield, 2),
                roundl(Fruit_Yield, 2),
                roundl(Number_of_marketable_fruit, 2),
                roundl(Average_marketable_fruit_weight, 2),
                roundl(Individual_Marketable_Fruit_Yield, 2),
                roundl(Marketable_Fruit_Yield, 2),
                roundl(Marketable_Fruit_Number_2Ws, 2),
                roundl(Marketable_Fruit_Weight_2Ws, 2),
                roundl(Individual_Marketable_Fruit_Yield_2Ws, 2),
                roundl(Marketable_Fruit_Yield_2Ws, 2),
            )
    return res


df_data, df_color = readExcelSheet(sheet)
avg = experimentCellsToNpArray(df_data, df_color)


def getAvgAllReplication(avg):
    tmp = avg.drop(columns=["Replication", "Treatment"], axis=1)
    res = pd.DataFrame(columns=tmp.columns)
    for (spacing, truss), group_data in tmp.groupby(["Spacing", "Truss"]):
        row = pd.Series()
        row["Spacing"] = spacing
        row["Truss"] = truss
        for col in group_data.columns:
            if col not in ["Spacing", "Truss"]:
                groups_mean = [x[0] for x in group_data[col]]
                groups_ssw = sum([x[1] for x in group_data[col]])
                groups_size = [x[2] for x in group_data[col]]

                mean, ssw, size = ANOVA.anova_table(
                    groups_mean, groups_ssw, groups_size
                )
                print(col, mean, ssw, size)
                row[col] = [mean, ssw, size]
        res = res._append(row, ignore_index=True)
    print(res)
    return res


tmp = getAvgAllReplication(avg)


# res_print = tmp.copy()
# res_print.columns = [
#     "Spacing",
#     "Truss",
#     # "Truss 1",
#     # "Truss 2",
#     # "Truss 3",
#     # "Truss 4",
#     # "Truss 5",
#     "Number of\nfruit",
#     "Average\nfruit weight",
#     "Individual\nFruit Yield",
#     "Fruit Yield",
#     "Number of\nmarketable fruit",
#     "Average marketable\nfruit weight",
#     "Individual Marketable\nFruit Yield",
#     "Marketable Fruit\nYield",
#     "Marketable Fruit\nNumber 2Ws",
#     "Marketable Fruit\nWeight 2Ws",
#     "Individual Marketable\nFruit Yield 2Ws",
#     "Marketable Fruit\nYield 2Ws",
# ]
# res_print["Spacing"] = res_print["Spacing"].apply(
#     lambda x: (
#         f"S{x}" if pd.notna(x) and pd.to_numeric(x, errors="coerce") is not None else x
#     )
# )
# res_print["Truss"] = res_print["Truss"].apply(
#     lambda x: (
#         f"T{x}" if pd.notna(x) and pd.to_numeric(x, errors="coerce") is not None else x
#     )
# )


# # Open the saved PNG image
import os
import StatisticsPNG as SPNG

# columns_to_remove = [2, 4, 6, 17, 18]
# df = df.drop(df.columns[columns_to_remove], axis=1)
# avg = experimentCellsToNpArray(df)
# res = assessAllReplications(avg)

SPNG.configTable(tmp, tableDir)
os.system("start " + tableDir)
